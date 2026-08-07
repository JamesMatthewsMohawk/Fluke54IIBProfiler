"""CSV/Excel export for one or more runs, shared by the Database and Profile tabs.

Each run is passed as (Run, [(elapsed_time_s, temperature_c), ...]) -- raw
Celsius, matching the DB's canonical storage. CSV always writes Celsius, for
data fidelity; Excel converts to the given display unit since it's meant to
be read alongside what's shown on screen. run_date is stored in UTC and
formatted per use_local_time -- see app/time_format.py.
"""
from __future__ import annotations

import csv

from openpyxl import Workbook
from openpyxl.chart import LineChart, Reference
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

from app.models import Run
from app.time_format import format_run_date
from app.units import convert_from_celsius

RunSeries = tuple[Run, list[tuple[float, float]]]


def export_runs_to_csv(path: str, runs: list[RunSeries], use_local_time: bool = False) -> None:
    with open(path, "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow([
            "run_date", "plant", "tunnel", "source_unit_at_log_time",
            "peak_temp_c", "min_temp_c", "elapsed_time_s", "temperature_c",
        ])
        for run, points in runs:
            run_date = format_run_date(run.run_date, use_local_time)
            for elapsed_time_s, temperature_c in points:
                writer.writerow([
                    run_date, run.plant, run.tunnel, run.source_unit,
                    f"{run.peak_temp_c:.2f}", f"{run.min_temp_c:.2f}",
                    elapsed_time_s, f"{temperature_c:.3f}",
                ])


def export_runs_to_excel(path: str, runs: list[RunSeries], unit: str, use_local_time: bool = False) -> None:
    wb = Workbook()
    data_ws = wb.active
    data_ws.title = "Profile Data"
    bold = Font(bold=True)

    chart = LineChart()
    chart.title = "Temperature Profile"
    chart.x_axis.title = "Elapsed Time (s)"
    chart.y_axis.title = f"Temperature (°{unit})"
    chart.style = 2

    col = 1
    for run, points in runs:
        label = f"{run.plant} / {run.tunnel}" if run.plant else run.tunnel
        run_date = format_run_date(run.run_date, use_local_time)
        data_ws.cell(row=1, column=col, value=f"{label} ({run_date})").font = bold
        data_ws.cell(row=2, column=col, value="Time (s)").font = bold
        data_ws.cell(row=2, column=col + 1, value=f"Temp (°{unit})").font = bold

        for i, (elapsed_time_s, temperature_c) in enumerate(points):
            data_ws.cell(row=3 + i, column=col, value=elapsed_time_s)
            data_ws.cell(row=3 + i, column=col + 1, value=round(convert_from_celsius(temperature_c, unit), 3))

        last_row = 2 + len(points)
        data_ref = Reference(data_ws, min_col=col + 1, min_row=2, max_row=last_row)
        cats_ref = Reference(data_ws, min_col=col, min_row=3, max_row=last_row)
        chart.add_data(data_ref, titles_from_data=True)
        chart.set_categories(cats_ref)

        col += 3  # one blank column of separation between each run's table

    data_ws.add_chart(chart, f"{get_column_letter(col + 1)}2")

    summary_ws = wb.create_sheet("Summary")
    headers = ["Run Date", "Plant", "Tunnel", f"Peak (°{unit})", f"Min (°{unit})", "Points", "Logged °"]
    for c, header in enumerate(headers, start=1):
        summary_ws.cell(row=1, column=c, value=header).font = bold
    for r, (run, _points) in enumerate(runs, start=2):
        peak = convert_from_celsius(run.peak_temp_c, unit)
        min_t = convert_from_celsius(run.min_temp_c, unit)
        summary_ws.cell(row=r, column=1, value=format_run_date(run.run_date, use_local_time))
        summary_ws.cell(row=r, column=2, value=run.plant)
        summary_ws.cell(row=r, column=3, value=run.tunnel)
        summary_ws.cell(row=r, column=4, value=round(peak, 2))
        summary_ws.cell(row=r, column=5, value=round(min_t, 2))
        summary_ws.cell(row=r, column=6, value=run.measurement_count)
        summary_ws.cell(row=r, column=7, value=run.source_unit)

    wb.save(path)
